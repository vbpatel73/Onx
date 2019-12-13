from seleniumrequests import Chrome
import time
from datetime import datetime
import openpyxl as xl
import asyncio
import random

wb = xl.load_workbook('C:\\Users\\patelvi\\Desktop\\sele_req.xlsx')
data = wb['Sheet1']
pat = []
for x in range(227,228):
    o_id = data.cell(x, 1).value
    o_code = data.cell(x, 2).value
    p_ln = data.cell(x, 5).value
    p_mn = data.cell(x, 6).value
    p_fn = data.cell(x, 7).value
    p_db = data.cell(x, 8).value
    p_sex = data.cell(x, 9).value
    p_hc = data.cell(x, 3).value
    p_version = data.cell(x, 4).value
    p_address = data.cell(x, 10).value
    p_city = data.cell(x, 11).value
    p_pcode = data.cell(x, 13).value
    p_ph = data.cell(x, 14).value

    d = {
        'Order.OrderId': o_id,
        'Order.OrderCode': o_code,
        'LockSeconds': '600',
        'Order.StatusId': '2',
        'RejectCode': '',
        'Isvalid': 'True',
        'Order.IsProviderFound': 'True',
        'Order.ClientID': '131313',
        'Order.CopytoClientID': '',
        'Order.LisMixedAddress': '',
        'Order.PatientSweeperMixedAddress': '',
        'Order.FitkitMailingSweeperMixedAddress': '',
        'Order.DefaultHomelessAddress': '505 University Ave|Toronto|ON|M5G2P1',
        'IsExistingHCN': 'False',
        'Order.IsLongTermFollowUp': '',
        'Order.RequesterTypeId': '1',
        'Order.MobileCoachId': '',
        'Order.CpsoCnoNumber': '99999',
        'Order.OhipBillNumber': '131313',
        'Order.LocationID': '',
        'Order.RequesterLastName': 'DR. IT TESTING',
        'Order.RequesterMiddleName': 'DR. IT TESTING',
        'Order.RequesterFirstName': 'DR. IT TESTING',
        'Order.RequesterOfficeAddress': '100 INTERNATIONAL BLVD',
        'Order.RequesterOfficeCity': 'TORONTO',
        'Order.RequesterOfficeProvince': 'ON',
        'Order.RequesterOfficePostalCode': 'M9W6J6',
        'Order.RequesterOfficePhoneNumber': '(416) 675-4530',
        'Order.RequesterOfficeFaxNumber': '(416) 213-1930',
        'Order.CopytoLastName': '',
        'Order.CopytoMiddleName': '',
        'Order.CopytoFirstName': '',
        'Order.CopytoOfficeAddress': '',
        'Order.CopytoOfficeCity': '',
        'Order.CopytoOfficeProvince': '',
        'Order.CopytoOfficePostalCode': '',
        'Order.CopytoOfficePhoneNumber': '',
        'Order.CopytoOfficeFaxNumber': '',
        'Order.PatientLastName': p_ln,
        'Order.PatientMiddleName': p_mn,
        'Order.PatientFirstName': p_fn,
        'Order.PatientDateOfBirth': p_db,  # 1952-05-18
        'Order.PatientOhipNumber': p_hc,
        'Order.PatientOhipVersion': p_version,
        'Order.PatientSexId': p_sex,  # 1,2
        'Order.PatientResponseCode': '50-Health card passed validation.',
        'Order.PatientPrimaryPhoneNumber': '(416)416-4160',#p_ph,
        'Order.PatientPhoneTypeId': '2',
        'Order.PatientAddress': p_address,
        'Order.isNoPatientAddress': 'false',
        'Order.PatientCity': p_city,
        'Order.PatientProvince': 'ON',
        'Order.PatientPostalCode': p_pcode,
        'Order.PatientPhoneExtension': '',
        'Order.PatientCellPhoneNumber': '',
        'Order.IsDifferentKitMailingAddress': 'false',
        'Order.FitkittMailingAddress': p_address,
        'Order.FitkittMailingFacilityName': '',
        'Order.FitkittMailingCity': p_city,
        'Order.FitkittMailingProvince': 'ON',
        'Order.FitkittMailingPostalCode': p_pcode,
        'Order.FitkittMailingPrimaryPhoneNumber':'(416)416-4160',# p_ph,
        'Order.FitkittMailingPhoneExtension': '',
        'Order.FitkittMailingPhoneTypeId': '2',
        'Order.RequesterSignature': 'True',
        'Order.PatientRequiresNewKit': 'false',
        'Order.Notes': ''}
    pat.append(d)

#print(pat[1])
user = [#"Vikash@lifelabsccofit.onmicrosoft.com",
        #"VikashPA@lifelabsccofit.onmicrosoft.com",
        #"Anubhav@lifelabsccofit.onmicrosoft.com",
        #"AnubhavAdmin@lifelabsccofit.onmicrosoft.com",
        #"AnubhavPA@lifelabsccofit.onmicrosoft.com",
        #"Testuser@lifelabsccofit.onmicrosoft.com",
        #"Mahnoosh@lifelabsccofit.onmicrosoft.com",
        #"Nadeem@lifelabsccofit.onmicrosoft.com",
        #"VikashAdmin@lifelabsccofit.onmicrosoft.com",
        #"Test@lifelabsccofit.onmicrosoft.com",
        'Tester1@lifelabsccofit.onmicrosoft.com',
        'Tester2@lifelabsccofit.onmicrosoft.com',
        'Tester3@lifelabsccofit.onmicrosoft.com',
        'Tester4@lifelabsccofit.onmicrosoft.com',
        'Tester5@lifelabsccofit.onmicrosoft.com',
        'Tester6@lifelabsccofit.onmicrosoft.com',
        'Tester7@lifelabsccofit.onmicrosoft.com',
        'Tester8@lifelabsccofit.onmicrosoft.com',
        'Tester9@lifelabsccofit.onmicrosoft.com',
        'Tester10@lifelabsccofit.onmicrosoft.com',
        'Tester11@lifelabsccofit.onmicrosoft.com',
        'Tester12@lifelabsccofit.onmicrosoft.com',
        'Tester13@lifelabsccofit.onmicrosoft.com',
        'Tester14@lifelabsccofit.onmicrosoft.com',
        'Tester15@lifelabsccofit.onmicrosoft.com',
        'Tester16@lifelabsccofit.onmicrosoft.com',
        'Tester17@lifelabsccofit.onmicrosoft.com',
        'Tester18@lifelabsccofit.onmicrosoft.com',
        'Tester19@lifelabsccofit.onmicrosoft.com',
        'Tester20@lifelabsccofit.onmicrosoft.com',
        'Tester21@lifelabsccofit.onmicrosoft.com',
        'Tester22@lifelabsccofit.onmicrosoft.com',
        'Tester23@lifelabsccofit.onmicrosoft.com',
        'Tester24@lifelabsccofit.onmicrosoft.com',
        'Tester25@lifelabsccofit.onmicrosoft.com',
        'Tester26@lifelabsccofit.onmicrosoft.com',
        'Tester27@lifelabsccofit.onmicrosoft.com',
        'Tester28@lifelabsccofit.onmicrosoft.com',
        'Tester29@lifelabsccofit.onmicrosoft.com',
        'Tester30@lifelabsccofit.onmicrosoft.com',
        'Tester31@lifelabsccofit.onmicrosoft.com',
        'Tester32@lifelabsccofit.onmicrosoft.com',
        'Tester33@lifelabsccofit.onmicrosoft.com',
        'Tester34@lifelabsccofit.onmicrosoft.com',
        'Tester35@lifelabsccofit.onmicrosoft.com',
        'Tester36@lifelabsccofit.onmicrosoft.com',
        'Tester37@lifelabsccofit.onmicrosoft.com',
        'Tester38@lifelabsccofit.onmicrosoft.com',
        'Tester39@lifelabsccofit.onmicrosoft.com',
        'Tester40@lifelabsccofit.onmicrosoft.com']

# Simple usage with built-in WebDrivers:
async def order(usr,patient):

    driver=Chrome("C:\\Users\\patelvi\\PycharmProjects\\NCE\\Common\\Drivers\\chromedriver.exe")
    driver.get("https://cco-fit-uat.lifelabs.com")
    time.sleep(1)
    #driver.maximize_window()
    driver.find_elements_by_name("loginfmt")[0].send_keys(usr)
    driver.find_elements_by_id("idSIButton9")[0].click()
    time.sleep(2)
    driver.find_elements_by_name("passwd")[0].send_keys("Shlok1998")
    time.sleep(2)
    driver.find_elements_by_id("idSIButton9")[0].click()
    time.sleep(3)
    driver.find_elements_by_id("idSIButton9")[0].click()
    time.sleep(2)
    #d_array.append(driver)
    await asyncio.sleep(0.1)
    print("--"+datetime.utcnow().strftime('%H:%M:%S.%f')[:-3])
    r1 = driver.request('POST', 'https://cco-fit-uat.lifelabs.com/Order/SaveKitting', data=patient)
    print(r1.text)
    '''
    r2 = driver.request('POST','https://cco-fit-uat.lifelabs.com/api/orders/get-lis-provider?searchText=313131&searchType=1')
    r3 = driver.request('POST','https://cco-fit-uat.lifelabs.com/api/orders/get-lis-patient?HCN=1001399029&HCNVersionCode=')
    r4 = driver.request('GET','https://cco-fit-uat.lifelabs.com/api/orders/get-hcn-orders?orderCode=14243E22&hcn=1001399177')
    print(r2.status_code)
    print(r2.text)
    print(r3.status_code)
    print(r4.text)
    '''
    #r1 = driver.request('POST', 'https://cco-fit-uat.lifelabs.com/Order/SaveKitting', data=patient)
    #print(datetime.utcnow().strftime('%H:%M:%S.%f')[:-3])
    #s1 = time.perf_counter()
    #r1 = driver.request('POST', 'https://cco-fit-uat.lifelabs.com/Order/SaveKitting', data=patient)
    #if r1.status_code != 200 :print(r1.status_code)
    #s2 = time.perf_counter()
    #print(round((s2 - s1),2))
    #driver.close()
    '''
    for x in range(15):
        response = driver.request('POST', 'https://cco-fit-uat.lifelabs.com/Order/SaveKitting', data=patient)
        print(response.text)

    driver.close()
    '''

o=[]
for i in range(3):
    usr=user[i]
    patient = pat[0]
    o.append(order(usr,patient))
    #print("---" + datetime.utcnow().strftime('%H:%M:%S.%f')[:-3])





async def main():
    # await asyncio.gather(*s)
    await asyncio.gather(*o)
    #print("----" + datetime.utcnow().strftime('%H:%M:%S.%f')[:-3])


if __name__ == "__main__":
    asyncio.run(main())
#time.sleep(1000)





































