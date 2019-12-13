import time , random
import asyncio
import requests as rq
import json
import openpyxl as xl


wb = xl.load_workbook("C:\\Users\\patelvi\\Desktop\\rest_api_cco.xlsx")
pat = wb['Sheet1']
acc=[]
ords=[]

for i in range(1,50):
    acc.append(str(pat.cell(i,2).value))
    ords.append(str(pat.cell(i,1).value))

#print(acc)

fit_url = "https://cco-fit-uat.lifelabs.com"
lis_url = "http://spwvapiwp101.diaglabs.corp.priv"  # "https://sdwvxldsk513.diaglabs.corp.priv"
token = "eyJhbGciOiJSUzI1NiIsImtpZCI6IjJhMTU0OTE1M2NjYzFiNmZkMWI3ZjgxZGQ2ZmQ4MGY5IiwidHlwIjoiSldUIn0.eyJuYmYiOjE1NDY0NDM3NzcsImV4cCI6MTU0NjQ1MDk3NywiaXNzIjoiaHR0cHM6Ly9zZHd2eGxkc2s1MTMuZGlhZ2xhYnMuY29ycC5wcml2OjgwOTUiLCJhdWQiOlsiaHR0cHM6Ly9zZHd2eGxkc2s1MTMuZGlhZ2xhYnMuY29ycC5wcml2OjgwOTUvcmVzb3VyY2VzIiwicHd0YXBpIl0sImNsaWVudF9pZCI6InB3dCIsInNjb3BlIjpbInB3dGFwaSJdfQ.xmlA-zY6UMIojfUsKpO7UcmGVuB8e419_hNpfsjwm6vQMruFx8vXwJ1wcyG9clYDLw5vuhT9snraBdGM2w9h7qgQixgsN8IlwFREBdpMn_xIXn5zmTWZEf3F5MYXED0U2jNJwJtyhnG79MwGmktqECN-xmVT2Yipi4AN2L2C1rxmitficO2pPe18T4LmYoOnUK6adI8S80sAKkMQLvMqTOSWmpHd4jwlASKK626v4cvtUR86RTYXya1i7mCAoy-YmOGt4CO4LrfebX5lslhmWt1uCeWzPO3qFm6SUBSjQL4sQ43I523gEx2rbT4dHzHJQl1HBCdtbjTj2LU6Sre9AQ"


async def generate_acc(order_code):
    data = {"orderCode": order_code}
    a = rq.post(lis_url + "/api/v2/FromCCOFitSrv/ccofit/generate-accession", data=json.dumps(data),
                headers={'Content-Type': 'application/json',
                         'Authorization': "Anything"})
    print(a.text)

async def hc_validation(hc, vc):
    d = {
        "hcn": hc,
        "hcnVersionCode": vc,
        "timeoutSeconds": 10
        }
    a = rq.post(lis_url + "/api/v2/FromCCOFitSrv/common/hcn-validation", data=json.dumps(d),
                headers={'Content-Type': 'application/json',
                         'Authorization': "Anything"})
    print(a.text)

async def get_patient(hc, vc):
    a = rq.get(lis_url + "/api/v2/FromCCOFitSrv/patients?hcn=" + str(hc) + "&hcnVersion=" + str(vc),
               headers={'Content-Type': 'application/json',
                        'Authorization': "Anything"})
    # b = json.loads(a.text)
    print(a.text)

async def get_provider(id, type):
    a = rq.get(lis_url + "/api/v1/FromCCOFitSrv/ccofit/providers?search=" + id + "&type=" + type,
               # type=LastName" ,/api/v2/FromCCOFitSrv
               headers={'Content-Type': 'application/json',
                        'Authorization': "Anything"})
    print(a.text)


async def set_coll_date(acc, order_id, coll_date):
    d = {
        "accession": acc,
        "orderCode": order_id,
        "sampleCollectionDate": coll_date,
        "kitReceiveLocation": "0011",
        "kitReceiveDate": "2018-11-14",
        "kitReceiveMethod": "M"
    }

    a = rq.put(fit_url + "/api/orders/set-sample-collection-date", data=json.dumps(d),
               headers={'Content-Type': 'application/json',
                        'Authorization': 'Bearer {}'.format(token)})
    print(str(a.status_code)+"|"+a.text)

    await time.sleep(0.01)

async def set_order_status(acc, order_id, status, date):
    d = {
        "accession": acc,
        "orderCode": order_id,
        "status": status,
        "isTestResultValid": "true",
        "resultedDate": date
    }

    a = rq.put(fit_url + "/api/orders/set-order-status", data=json.dumps(d),
               headers={'Content-Type': 'application/json',
                        'Authorization': 'Bearer {}'.format(token)})
    # b = a.text
    print(str(a.status_code) + "|" + a.text)
    await time.sleep(0.01)


async def set_lot_expiry(lot, expiry, acc, order_id):
    d = [{
        "lot": lot,
        "expiry": expiry,
        "mailingDate": "2019-12-12 14:20:25",
        "accession": acc,
        "orderCode": order_id
    }]

    a = rq.put(fit_url + "/api/orders/set-lot-and-expiry-for-kit", data=json.dumps(d),
               headers={'Content-Type': 'application/json',
                        'Authorization': 'Bearer {}'.format(token)})
    # b = a.text
    print(str(a.status_code) + "|" + a.text)
    await time.sleep(0.01)

async def get_order(order_id, acc):
    a = rq.get(fit_url + "/api/orders/orders?orderCode=" + order_id + "&accessionNumber=" + acc,
               headers={'Content-Type': 'application/json',
                        'Authorization': 'Bearer {}'.format(token)})
    print(str(a.status_code) + "|" + a.text)
    await time.sleep(0.01)

async def check_eligibility(hc, vc):
    data = {
        "hcn": hc,
        "hcnVersionCode": "BL"
    }
    a = rq.get(
        "http://spwvapiwp101.diaglabs.corp.priv//api//v2//FromCCOFitSrv/patients?hcn=" + str(hc) + "&hcnVersion=" + str(
            vc),  # data=json.dumps (data) ,
        headers={'Content-Type': 'application/json',
                 'Authorization': "Anything"})
    b = json.loads(a.text)
    c = b["isValid"]
    d = b["responseCode"]
    e = b["eligibilityCode"]
    # return (c,d,e)
    print(a.status_code)
    print(a.text)


async def reject_order(order_id):
    d = {"accessionNumber": order_id}
    a = rq.post("http://spwvapiwp101.diaglabs.corp.priv/api/v2/FromCCOFitSrv/ccofit/reject-order", data=json.dumps(d),
                headers={'Content-Type': 'application/json',
                         'Authorization': "Anything"})
    # print(d)
    print(a.text)
    print(a.status_code)

o=[]
for j in range(49) :
    ac =acc[j]
    ord = ords[j]

    o.append(set_lot_expiry("1234","2019-10-20",ac,ord))
    o.append(set_coll_date(ac,ord,"2019-11-11"))
    o.append(set_order_status(ac,ord,"6","2019-09-16"))
    o.append(get_order(ord,ac))





async def main():
    await asyncio.gather(*o)
    #await asyncio.gather(*problem_ords)
    #await asyncio.gather(*search_ords)

if __name__ == "__main__":
    asyncio.run(main())
#time.sleep(1000)
