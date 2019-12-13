from selenium import webdriver as wd
import time
from datetime import datetime
import openpyxl as xl
import asyncio
import requests as rq

'''
wb = xl.load_workbook('C:\\Users\\patelvi\\Desktop\\sele_req.xlsx')
data = wb['Sheet1']
pat = []
for x in range(221,226):
    o_id = data.cell(x, 1).value
    o_code = data.cell(x, 2).value
    p_ln = data.cell(x, 5).value
    p_mn = data.cell(x, 6).value
    p_fn = data.cell(x, 7).value
    p_db = data.cell(x, 8).value
    p_sex = data.cell(x, 9).value
    p_hc = data.cell(x, 3).value
    p_version = data.cell(x, 4).value
    p_address = data.cell(x, 10).value
    p_city = data.cell(x, 11).value
    p_pcode = data.cell(x, 13).value
    p_ph = data.cell(x, 14).value

    d = {
        'Order.OrderId': o_id,
        'Order.OrderCode': o_code,
        'LockSeconds': '600',
        'Order.StatusId': '2',
        'RejectCode': '',
        'Isvalid': 'True',
        'Order.IsProviderFound': 'True',
        'Order.ClientID': '131313',
        'Order.CopytoClientID': '',
        'Order.LisMixedAddress': '',
        'Order.PatientSweeperMixedAddress': '',
        'Order.FitkitMailingSweeperMixedAddress': '',
        'Order.DefaultHomelessAddress': '505 University Ave|Toronto|ON|M5G2P1',
        'IsExistingHCN': 'False',
        'Order.IsLongTermFollowUp': '',
        'Order.RequesterTypeId': '1',
        'Order.MobileCoachId': '',
        'Order.CpsoCnoNumber': '99999',
        'Order.OhipBillNumber': '131313',
        'Order.LocationID': '',
        'Order.RequesterLastName': 'DR. IT TESTING',
        'Order.RequesterMiddleName': 'DR. IT TESTING',
        'Order.RequesterFirstName': 'DR. IT TESTING',
        'Order.RequesterOfficeAddress': '100 INTERNATIONAL BLVD',
        'Order.RequesterOfficeCity': 'TORONTO',
        'Order.RequesterOfficeProvince': 'ON',
        'Order.RequesterOfficePostalCode': 'M9W6J6',
        'Order.RequesterOfficePhoneNumber': '(416) 675-4530',
        'Order.RequesterOfficeFaxNumber': '(416) 213-1930',
        'Order.CopytoLastName': '',
        'Order.CopytoMiddleName': '',
        'Order.CopytoFirstName': '',
        'Order.CopytoOfficeAddress': '',
        'Order.CopytoOfficeCity': '',
        'Order.CopytoOfficeProvince': '',
        'Order.CopytoOfficePostalCode': '',
        'Order.CopytoOfficePhoneNumber': '',
        'Order.CopytoOfficeFaxNumber': '',
        'Order.PatientLastName': p_ln,
        'Order.PatientMiddleName': p_mn,
        'Order.PatientFirstName': p_fn,
        'Order.PatientDateOfBirth': p_db,  # 1952-05-18
        'Order.PatientOhipNumber': p_hc,
       'Order.PatientOhipVersion': p_version,
        'Order.PatientSexId': p_sex,  # 1,2
        'Order.PatientResponseCode': '50-Health card passed validation.',
        'Order.PatientPrimaryPhoneNumber': '(416)416-4160',#p_ph,
        'Order.PatientPhoneTypeId': '2',
        'Order.PatientAddress': p_address,
        'Order.isNoPatientAddress': 'false',
        'Order.PatientCity': p_city,
        'Order.PatientProvince': 'ON',
        'Order.PatientPostalCode': p_pcode,
        'Order.PatientPhoneExtension': '',
        'Order.PatientCellPhoneNumber': '',
        'Order.IsDifferentKitMailingAddress': 'false',
        'Order.FitkittMailingAddress': p_address,
        'Order.FitkittMailingFacilityName': '',
        'Order.FitkittMailingCity': p_city,
        'Order.FitkittMailingProvince': 'ON',
        'Order.FitkittMailingPostalCode': p_pcode,
        'Order.FitkittMailingPrimaryPhoneNumber':'(416)416-4160',# p_ph,
        'Order.FitkittMailingPhoneExtension': '',
        'Order.FitkittMailingPhoneTypeId': '2',
        'Order.RequesterSignature': 'True',
        'Order.PatientRequiresNewKit': 'false',
        'Order.Notes': ''}
    pat.append(d)

#print(pat[1])
user = [#"Vikash@lifelabsccofit.onmicrosoft.com",
        #"VikashPA@lifelabsccofit.onmicrosoft.com",
        #"Anubhav@lifelabsccofit.onmicrosoft.com",
        #"AnubhavAdmin@lifelabsccofit.onmicrosoft.com",
        #"AnubhavPA@lifelabsccofit.onmicrosoft.com",
        #"Testuser@lifelabsccofit.onmicrosoft.com",
        #"Mahnoosh@lifelabsccofit.onmicrosoft.com",
        #"Nadeem@lifelabsccofit.onmicrosoft.com",
        #"VikashAdmin@lifelabsccofit.onmicrosoft.com",
        #"Test@lifelabsccofit.onmicrosoft.com",
        'Tester1@lifelabsccofit.onmicrosoft.com',
        'Tester2@lifelabsccofit.onmicrosoft.com',
        'Tester3@lifelabsccofit.onmicrosoft.com',
        'Tester4@lifelabsccofit.onmicrosoft.com',
        'Tester5@lifelabsccofit.onmicrosoft.com',
        'Tester6@lifelabsccofit.onmicrosoft.com',
        'Tester7@lifelabsccofit.onmicrosoft.com',
        'Tester8@lifelabsccofit.onmicrosoft.com',
        'Tester9@lifelabsccofit.onmicrosoft.com',
        'Tester10@lifelabsccofit.onmicrosoft.com',
        'Tester11@lifelabsccofit.onmicrosoft.com',
        'Tester12@lifelabsccofit.onmicrosoft.com',
        'Tester13@lifelabsccofit.onmicrosoft.com',
        'Tester14@lifelabsccofit.onmicrosoft.com',
        'Tester15@lifelabsccofit.onmicrosoft.com',
        'Tester16@lifelabsccofit.onmicrosoft.com',
        'Tester17@lifelabsccofit.onmicrosoft.com',
        'Tester18@lifelabsccofit.onmicrosoft.com',
        'Tester19@lifelabsccofit.onmicrosoft.com',
        'Tester20@lifelabsccofit.onmicrosoft.com',
        'Tester21@lifelabsccofit.onmicrosoft.com',
        'Tester22@lifelabsccofit.onmicrosoft.com',
        'Tester23@lifelabsccofit.onmicrosoft.com',
        'Tester24@lifelabsccofit.onmicrosoft.com',
        'Tester25@lifelabsccofit.onmicrosoft.com',
        'Tester26@lifelabsccofit.onmicrosoft.com',
        'Tester27@lifelabsccofit.onmicrosoft.com',
       'Tester28@lifelabsccofit.onmicrosoft.com',
        'Tester29@lifelabsccofit.onmicrosoft.com',
        'Tester30@lifelabsccofit.onmicrosoft.com',
        'Tester31@lifelabsccofit.onmicrosoft.com',
        'Tester32@lifelabsccofit.onmicrosoft.com',
        'Tester33@lifelabsccofit.onmicrosoft.com',
        'Tester34@lifelabsccofit.onmicrosoft.com',
        'Tester35@lifelabsccofit.onmicrosoft.com',
        'Tester36@lifelabsccofit.onmicrosoft.com',
        'Tester37@lifelabsccofit.onmicrosoft.com',
        'Tester38@lifelabsccofit.onmicrosoft.com',
        'Tester39@lifelabsccofit.onmicrosoft.com',
        'Tester40@lifelabsccofit.onmicrosoft.com']
'''
fit_url ="https://cco-fit-uat.lifelabs.com"
lis_url = "http://spwvapiwp101.diaglabs.corp.priv" #"https://sdwvxldsk513.diaglabs.corp.priv"
token ="eyJhbGciOiJSUzI1NiIsImtpZCI6IjJhMTU0OTE1M2NjYzFiNmZkMWI3ZjgxZGQ2ZmQ4MGY5IiwidHlwIjoiSldUIn0.eyJuYmYiOjE1NDY0NDM3NzcsImV4cCI6MTU0NjQ1MDk3NywiaXNzIjoiaHR0cHM6Ly9zZHd2eGxkc2s1MTMuZGlhZ2xhYnMuY29ycC5wcml2OjgwOTUiLCJhdWQiOlsiaHR0cHM6Ly9zZHd2eGxkc2s1MTMuZGlhZ2xhYnMuY29ycC5wcml2OjgwOTUvcmVzb3VyY2VzIiwicHd0YXBpIl0sImNsaWVudF9pZCI6InB3dCIsInNjb3BlIjpbInB3dGFwaSJdfQ.xmlA-zY6UMIojfUsKpO7UcmGVuB8e419_hNpfsjwm6vQMruFx8vXwJ1wcyG9clYDLw5vuhT9snraBdGM2w9h7qgQixgsN8IlwFREBdpMn_xIXn5zmTWZEf3F5MYXED0U2jNJwJtyhnG79MwGmktqECN-xmVT2Yipi4AN2L2C1rxmitficO2pPe18T4LmYoOnUK6adI8S80sAKkMQLvMqTOSWmpHd4jwlASKK626v4cvtUR86RTYXya1i7mCAoy-YmOGt4CO4LrfebX5lslhmWt1uCeWzPO3qFm6SUBSjQL4sQ43I523gEx2rbT4dHzHJQl1HBCdtbjTj2LU6Sre9AQ"


def get_patient ( hc , vc ):
   a = rq.get (lis_url + "/api/v2/FromCCOFitSrv/patients?hcn=" + str (hc) + "&hcnVersion=" + str (vc) ,
               headers={'Content-Type': 'application/json' ,
                        'Authorization': "Anything"})
   # b = json.loads(a.text)
   print (a.text)


# return (a.text,a.status_code)


# return (a.text,a.status_code)

def get_provider ( id , type ):
   # a = rq.get (lis_url + "/api/v1/FromCCOFitSrv/ccofit/providers?search="+id+"&type=BillingNumber",headers = {'Content-Type':'application/json',
   #                               'Authorization':"Anything"})
   a = rq.get (lis_url + "/api/v2/FromCCOFitSrv/ccofit/providers?search=" + id + "&type=" + type ,
               # type=LastName" ,/api/v2/FromCCOFitSrv
               headers={'Content-Type': 'application/json' ,
                        'Authorization': "Anything"})
   print (a.text)


# Simple usage with built-in WebDrivers:
async def order():
   '''
   driver = wd.Chrome("C:\\Users\\patelvi\\PycharmProjects\\NCE\\Common\\Drivers\\chromedriver.exe")
   driver.implicitly_wait(3)

   driver.get("https://cco-fit-uat.lifelabs.com")
   time.sleep(1)
   driver.maximize_window()
   driver.find_elements_by_name("loginfmt")[0].send_keys(usr)#"Tester1@lifelabsccofit.onmicrosoft.com")
   time.sleep(2)
   driver.find_elements_by_id("idSIButton9")[0].click()
   time.sleep(2)
   driver.find_elements_by_name("passwd")[0].send_keys("Shlok1998")
   time.sleep(2)
   driver.find_elements_by_id("idSIButton9")[0].click()
   time.sleep(2)
   driver.find_elements_by_id("idSIButton9")[0].click()

   ck = driver.get_cookie('.AspNet.Cookies')
   #print(ck['value'])
   #print(ck['name'])
   a = rq.session()
   print(datetime.utcnow().strftime('%H:%M:%S.%f')[:-3])
   for name in ck:
      a.cookies.set(ck['name'], ck['value'])
   url = "https://cco-fit-uat.lifelabs.com/Search/GetSearchOrderList?sort=&page=1&pageSize=10&group=&filter=&SearchText=1015423476&searchType=2"
   r1 = rq.post(url, cookies=a.cookies)
   # r2 = rq.post('https://cco-fit-uat.lifelabs.com/search/get-lis-provider?searchText=313131&searchType=1',cookies=a.cookies)
   # r3 = rq.post('https://cco-fit-uat.lifelabs.com/search/get-lis-patient?HCN=1001399029&HCNVersionCode=',cookies=a.cookies)
   r4 = rq.get('https://cco-fit-uat.lifelabs.com/api/orders/get-hcn-orders?orderCode=14243E22&hcn=1001399177',
               cookies=a.cookies)
   print(r1.text)

   print(r4.text)
   '''
   print(datetime.utcnow().strftime('%H:%M:%S.%f')[:-3])
   get_patient('1015423476','')
   get_provider('131313','BillingNumber')
o=[]
for i in range(3000):
    #usr=user[i]
    #patient = pat[i]
    o.append(order())
    #print("---" + datetime.utcnow().strftime('%H:%M:%S.%f')[:-3])





async def main():
    # await asyncio.gather(*s)
    await asyncio.gather(*o)
    #print("----" + datetime.utcnow().strftime('%H:%M:%S.%f')[:-3])


if __name__ == "__main__":
    asyncio.run(main())
#time.sleep(1000)

